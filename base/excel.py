"""Módulo de manipulação de planilhas Excel
"""

# Third Part imports
# from openpyxl import load_workbook
# import xlsxwriter
import os
import platform
import xlsxwriter
import time
import openpyxl
import getpass

# Own Imports
from base.comandos import clear_screen


class ExcelStuffs:
    """Classe para manipular objetos Excel
    """

    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), "configs.xlsx")

    def checar_plan(self, caminho: str = caminho):
        """Checar existenta da planilha de configurações

        Args:
            caminho (str, optional): Path para a planilhas.
        """
        clear_screen()
        if not os.path.exists(caminho):
            print('Planilha inexistente.')
            self.criar_plan()


    def criar_plan(self, caminho: str = caminho):
        """Criar planilha

        Args:
            caminho (str, optional): Path para a planilha.
        """
        clear_screen()
        print("Criando Planilha")
        workbook = xlsxwriter.Workbook(caminho)
        worksheet = workbook.add_worksheet('Planilha1')
        worksheet.write(0, 0, 'Destinatários')
        worksheet.write(0, 1, 'Login Valor')
        worksheet.write(0, 2, 'Senha Valor')
        worksheet.write(0, 3, 'Lista de clientes')
        worksheet.write(0, 4, 'Remetente')
        worksheet.write(0, 5, 'Senha remetente')
        workbook.close()
        input("Aperte Enter para abrir a planilha para preencher os dados das configurações," +
              " salvar e fechar a planilha: ")
        clear_screen()

        # Abrindo planilha
        if platform.system() == 'Linux':
            os.system('xdg-open ' + f'"{caminho}"')
        else:
            os.system('open ' + f'"{caminho}"')
        clear_screen()
        time.sleep(2)
        input("Insira as informações, salve o arquivo, feche-o e aperte qualquer teclar: ")

    def coletando_dados(self, caminho: str = caminho):
        """Função que coletará dados na planilhas de configs.

        Args:
            caminho (str, optional): Path da planilhas configs.xlsx.

        Returns:
            Dict: Retorna um dicionário com os dados da planilha de configs.
        """

        # Jogando Excel no objeto

        wb = openpyxl.load_workbook(filename=caminho, read_only=True)
        # Coletando e-mails da lista
        lista_email = []
        for i in range(2, 500):
            if wb['Planilha1']["A{}".format(i)].value is None:
                break
            else:
                lista_email.append(wb['Planilha1']["A{}".format(i)].value + ',')

        # Coletando clientes da lista
        clientes = []
        for i in range(2, 500):
            if wb['Planilha1']["D{}".format(i)].value is None:
                break
            else:
                clientes.append(wb['Planilha1']["D{}".format(i)].value)

        # E-mail
        sender_email = wb['Planilha1']["E2"].value

        # Senha

        password_email = wb['Planilha1']["F2"].value
        if password_email is None:
            password_email = getpass.getpass()

        # Login VALOR

        loginvalor = wb['Planilha1']["B2"].value

        # Senha VALOR

        senhavalor = wb['Planilha1']["C2"].value
        if senhavalor is None:
            senhavalor = getpass.getpass('')

        wb.close()

        return {"emails": lista_email, "clientes": clientes, "remetente": sender_email, "senha email": password_email,
                "loginvalor": loginvalor, "senha valor": senhavalor}


if __name__ == "__main__":
    ExcelStuffs.checar_plan()
